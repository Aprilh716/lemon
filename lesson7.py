import openpyxl
import requests
import pprint

#读取数据
def readData(filName,sheetName):
    wx= openpyxl.open(filename=filName)
    sheet1=wx[sheetName]
    requestList=[]
    for i in range(2,sheet1.max_row+1):
        case=dict(
                  case_id=sheet1.cell(row=i, column=1).value,
                  url=sheet1.cell(row=i,column=5).value,
                  data=sheet1.cell(row=i,column=6).value,
                  expected=sheet1.cell(row=i,column=7).value
                  )
        # print(case)
        requestList.append(case)
    # print(requestList)
    return requestList #返回读取到的数据用于接口请求

#发送requests请求函数
header = {'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json'}
def apiRequest(url,data,headers=header):
    response=requests.post(url=url, json=data, headers=headers)
    res=response.json()
    # pprint.pprint(res)
    return res

#回写测试执行的结果
def writeResult(filename,sheetname,row,column,result):
    wx1=openpyxl.open(filename=filename)
    sheet2=wx1[sheetname]
    sheet2.cell(row,column).value=result
    wx1.save(filename)

#执行测试用例
token = ''
def exec_case(filename,sheetname):
    allDate=readData(filName=filename,sheetName=sheetname)
    for i in allDate:
        id=i['case_id']
        url=i['url']
        data=eval(i['data'])
        #获取token
        if sheetname=='login' and id==1:
            respon = apiRequest(url=url, data=data)
            token1=respon['data']['token_info']['token']
            global token
            token=token1
        #判断请求接口是否要传token
        if sheetname=='register' or sheetname=='login':
            response1=apiRequest(url=url,data=data)
        else:
            headers = {'X-Lemonban-Media-Type': 'lemonban.v2',
                      'Content-Type': 'application/json',
                      'Authorization': 'Bearer ' + token
                      }
            response1 = apiRequest(url=url, data=data,headers=headers)
            print(response1,'AAA')
        excepted_msg=eval(i['expected'])['msg']
        real_msg=response1['msg']
        if(excepted_msg==real_msg):
            print(f"用例{id}测试执行通过")
            writeResult(filename=filename,sheetname=sheetname,row=id+1,column=8,result='通过')
        else:
            print(f"用例{id}测试执行不通过")
            writeResult(filename=filename, sheetname=sheetname, row=id+1, column=8, result='不通过')
#执行注册测试用例
exec_case('test_case_api.xlsx','register')
# # #执行登录测试用例
exec_case('test_case_api.xlsx','login')
#执行充值测试用例
# exec_case('test_case_api.xlsx','recharge')